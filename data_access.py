"""
Data access layer — Google Sheets API edition (reads EVERYTHING).

Instead of one published-CSV link per tab (which can only ever return a single
tab, as-published), this talks to the Google Sheets API. Given an API key and
one or more spreadsheet IDs, it automatically:
  - discovers EVERY tab in each spreadsheet, and
  - reads ALL columns/rows of each tab.
Add more files later by adding their IDs to SPREADSHEET_IDS — no code change.

------------------------------------------------------------------------------
SETUP  (all in Render -> your service -> Environment)
------------------------------------------------------------------------------
    GOOGLE_API_KEY   = <a Google API key with the "Google Sheets API" enabled>
    SPREADSHEET_IDS  = <id1>,<id2>,...   (comma-separated; one or many files)

The spreadsheet ID is the long id in the NORMAL edit URL:
    https://docs.google.com/spreadsheets/d/<THIS_IS_THE_ID>/edit
NOTE: it is NOT the "/d/e/2PACX-....../pub" published token — that won't work
with the API. Open the sheet normally and copy the id between /d/ and /edit.

The spreadsheet must be readable by the key: set Share -> General access ->
"Anyone with the link" -> Viewer.

------------------------------------------------------------------------------
MERGED CELLS
------------------------------------------------------------------------------
The plain `/values/` endpoint returns a merged cell's text ONLY in the merge's
top-left ("anchor") cell; every other cell the merge covers comes back blank.
For manufacturing sheets that merge a Client / Job / PO across the several
component rows beneath it, that means those rows lose their key values — so
filtering, searching and counting are wrong, and continuation rows can even be
dropped as "blank".

To fix this we read the tab with `includeGridData=true`, which also returns the
sheet's `merges` metadata, and we PROPAGATE each merge's anchor value across
every cell in its range BEFORE building rows. See `_api_values`.

------------------------------------------------------------------------------
FALLBACK
------------------------------------------------------------------------------
If GOOGLE_API_KEY / SPREADSHEET_IDS are not set, this falls back to the older
behavior: any env var whose value is a published-CSV link (contains
"output=csv") is treated as one sheet. (One tab each — the limited mode.)
"""

import os
import io
import csv
import json
import time
import urllib.parse
import urllib.request
import urllib.error


GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "").strip()
_CSV_MARKER = "output=csv"
_CACHE_TTL = 30                       # seconds
_cache: dict[str, tuple[float, list]] = {}        # source-key -> (ts, rows)
_reg_cache: dict[str, object] = {"ts": 0.0, "reg": None}


# ---------------------------------------------------------------------
# LOW-LEVEL: Google Sheets API
# ---------------------------------------------------------------------
def _api_get(url: str) -> dict:
    req = urllib.request.Request(url, headers={"Accept": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", "replace")[:400]
        raise RuntimeError(f"Google Sheets API HTTP {e.code}: {body}") from e


def _api_meta(spreadsheet_id: str) -> dict:
    sid = urllib.parse.quote(spreadsheet_id)
    return _api_get(f"https://sheets.googleapis.com/v4/spreadsheets/{sid}?key={GOOGLE_API_KEY}")


def _dedup_headers(raw: list) -> list[str]:
    """Turn a raw header row into unique, non-empty column names.

    Blank cells become col1/col2/... (by position, matching the old behavior);
    duplicates — e.g. a horizontally merged group header propagated across the
    columns it spans — get a ' (2)', ' (3)' suffix so they don't collapse into
    a single dict key and silently drop columns.
    """
    headers: list[str] = []
    counts: dict[str, int] = {}
    for i, h in enumerate(raw):
        name = (str(h).strip() if h is not None else "") or f"col{i + 1}"
        if name in counts:
            counts[name] += 1
            name = f"{name} ({counts[name]})"
        else:
            counts[name] = 1
        headers.append(name)
    return headers


def _api_values(spreadsheet_id: str, tab: str) -> list[dict]:
    """Read a tab as row dicts, correctly expanding merged cells.

    Uses includeGridData=true so we get BOTH the grid values and the tab's
    `merges` metadata in a single call, then copies every merge's anchor value
    into all the cells it covers before the header row and the data rows are
    built. This is what makes vertically-merged Client/Job/PO columns show up
    on every row instead of only the first.
    """
    sid = urllib.parse.quote(spreadsheet_id)
    # Quote the tab title so titles with spaces/symbols work as a range.
    rng = urllib.parse.quote(f"'{tab.replace(chr(39), chr(39) * 2)}'")
    fields = urllib.parse.quote(
        "sheets(merges,data(startRow,startColumn,rowData.values.formattedValue))"
    )
    url = (f"https://sheets.googleapis.com/v4/spreadsheets/{sid}"
           f"?ranges={rng}&includeGridData=true&fields={fields}"
           f"&key={GOOGLE_API_KEY}")
    resp = _api_get(url)

    sheets = resp.get("sheets") or []
    if not sheets:
        return []
    sheet0 = sheets[0]
    data = sheet0.get("data") or []
    if not data:
        return []
    block = data[0]
    row_off = block.get("startRow", 0) or 0        # 0 when reading the whole tab
    col_off = block.get("startColumn", 0) or 0

    # Build a rectangular grid of displayed strings (None for empty cells).
    grid: list[list] = []
    for rd in (block.get("rowData") or []):
        cells = rd.get("values") or []
        grid.append([(c.get("formattedValue") if c else None) for c in cells])
    if not grid:
        return []

    width = max((len(r) for r in grid), default=0)
    if width == 0:
        return []
    for r in grid:
        if len(r) < width:
            r.extend([None] * (width - len(r)))

    # --- KEY FIX: propagate every merged cell's anchor value across its range.
    for m in (sheet0.get("merges") or []):
        r0 = m.get("startRowIndex", 0) - row_off
        r1 = m.get("endRowIndex", 0) - row_off          # exclusive
        c0 = m.get("startColumnIndex", 0) - col_off
        c1 = m.get("endColumnIndex", 0) - col_off        # exclusive
        if r0 < 0 or c0 < 0 or r0 >= len(grid) or c0 >= width:
            continue
        anchor = grid[r0][c0]
        if anchor in (None, ""):
            continue
        for rr in range(r0, min(r1, len(grid))):
            row = grid[rr]
            for cc in range(c0, min(c1, width)):
                if row[cc] in (None, ""):
                    row[cc] = anchor

    headers = _dedup_headers(grid[0])
    out = []
    for r in grid[1:]:
        if not any((str(c).strip() if c is not None else "") for c in r):
            continue
        out.append({
            headers[i]: (str(r[i]).strip() if i < len(r) and r[i] is not None else "")
            for i in range(len(headers))
        })
    return out


# ---------------------------------------------------------------------
# SHEET REGISTRY  (name -> source descriptor)
#   ("api", spreadsheet_id, tab_title)  |  ("csv", url, "")  |  ("error", msg, "")
# ---------------------------------------------------------------------
def _build_registry() -> dict[str, tuple]:
    reg: dict[str, tuple] = {}
    ids = [s.strip() for s in os.environ.get("SPREADSHEET_IDS", "").split(",") if s.strip()]

    if GOOGLE_API_KEY and ids:
        for sid in ids:
            try:
                meta = _api_meta(sid)
            except Exception as e:   # noqa: BLE001 - surface per-file, keep going
                reg[f"[error:{sid[:10]}]"] = ("error", str(e), "")
                continue
            stitle = (meta.get("properties") or {}).get("title") or sid[:10]
            for sh in meta.get("sheets", []):
                tab = (sh.get("properties") or {}).get("title")
                if not tab:
                    continue
                name = tab if tab not in reg else f"{stitle}::{tab}"
                reg[name] = ("api", sid, tab)
        return reg

    # ---- fallback: published-CSV links in env vars (one tab each) ----
    explicit = os.environ.get("SHEETS", "").strip()
    if explicit:
        for n in [x.strip() for x in explicit.split(",") if x.strip()]:
            v = os.environ.get(n, "").strip()
            if v:
                reg[n] = ("csv", v, "")
        return reg

    for name, val in os.environ.items():
        v = (val or "").strip()
        if _CSV_MARKER in v or v.lower().endswith((".csv", ".xlsx")):
            reg[name] = ("csv", v, "")
    return reg


def _registry() -> dict[str, tuple]:
    now = time.time()
    if _reg_cache["reg"] is not None and (now - _reg_cache["ts"]) < _CACHE_TTL:
        return _reg_cache["reg"]        # type: ignore[return-value]
    reg = _build_registry()
    _reg_cache["reg"] = reg
    _reg_cache["ts"] = now
    return reg


def list_sheet_names() -> list[str]:
    return sorted(_registry().keys())


def _resolve(sheet: str) -> tuple:
    reg = _registry()
    if sheet not in reg:
        raise KeyError(sheet)
    return reg[sheet]


# ---------------------------------------------------------------------
# LOADING (cached, source-agnostic)
# ---------------------------------------------------------------------
def _load(source: tuple) -> list[dict]:
    key = str(source)
    now = time.time()
    hit = _cache.get(key)
    if hit and (now - hit[0]) < _CACHE_TTL:
        return hit[1]

    kind = source[0]
    if kind == "api":
        rows = _api_values(source[1], source[2])
    elif kind == "csv":
        url = source[1]
        if url.startswith(("http://", "https://")):
            rows = _load_csv_url(url)
        elif url.lower().endswith(".xlsx"):
            rows = _load_xlsx(url)
        else:
            rows = _load_csv_file(url)
    elif kind == "error":
        raise RuntimeError(source[1])
    else:
        raise RuntimeError(f"Unknown source: {source!r}")

    _cache[key] = (now, rows)
    return rows


def _parse_csv(raw: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(raw))
    return [
        {(k or "").strip(): ("" if v is None else str(v).strip()) for k, v in row.items()}
        for row in reader
    ]


def _load_csv_url(url: str) -> list[dict]:
    req = urllib.request.Request(url, headers={"User-Agent": "manufacturing-mcp/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8-sig", errors="replace")
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Could not fetch CSV (HTTP {e.code}).") from e
    return _parse_csv(raw)


def _load_csv_file(path: str) -> list[dict]:
    with open(path, encoding="utf-8-sig") as f:
        return _parse_csv(f.read())


def _load_xlsx(path: str) -> list[dict]:
    from openpyxl import load_workbook
    from openpyxl.utils import range_boundaries
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active

    # Read the whole grid so we can expand merges the same way as the API path.
    grid = [list(row) for row in ws.iter_rows(values_only=True)]
    if not grid:
        return []
    width = max((len(r) for r in grid), default=0)
    for r in grid:
        if len(r) < width:
            r.extend([None] * (width - len(r)))

    for rng in list(ws.merged_cells.ranges):
        c0, r0, c1, r1 = range_boundaries(str(rng))   # 1-based, inclusive
        anchor = grid[r0 - 1][c0 - 1]
        if anchor in (None, ""):
            continue
        for rr in range(r0 - 1, r1):
            for cc in range(c0 - 1, c1):
                if grid[rr][cc] in (None, ""):
                    grid[rr][cc] = anchor

    headers = _dedup_headers([("" if h is None else str(h).strip()) for h in grid[0]])
    out = []
    for r in grid[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({
            headers[i]: ("" if i >= len(r) or r[i] is None else str(r[i]).strip())
            for i in range(len(headers))
        })
    return out


# ---------------------------------------------------------------------
# GENERIC QUERY API  (called by server.py — unchanged signatures)
# ---------------------------------------------------------------------
def _cell(row: dict, col: str):
    if col in row:
        return row[col]
    low = {k.lower(): v for k, v in row.items()}
    return low.get(str(col).lower())


def list_sheets() -> list[dict]:
    """Every tab across every connected spreadsheet, with columns + row count."""
    out = []
    for name in list_sheet_names():
        try:
            rows = _load(_resolve(name))
            out.append({
                "sheet": name,
                "row_count": len(rows),
                "columns": list(rows[0].keys()) if rows else [],
            })
        except Exception as e:   # noqa: BLE001
            out.append({"sheet": name, "error": str(e)})
    return out


def describe_sheet(sheet: str) -> dict:
    rows = _load(_resolve(sheet))
    return {
        "sheet": sheet,
        "columns": list(rows[0].keys()) if rows else [],
        "row_count": len(rows),
        "sample_rows": rows[:3],
    }


def get_sheet_rows(sheet: str, filters: dict | None = None, limit: int | None = None) -> list[dict]:
    rows = _load(_resolve(sheet))
    if filters:
        def match(r):
            for col, val in filters.items():
                if val in (None, ""):
                    continue
                if str(val).lower() not in str(_cell(r, col) or "").lower():
                    return False
            return True
        rows = [r for r in rows if match(r)]
    if limit and limit > 0:
        rows = rows[:limit]
    return rows


def search_sheet(sheet: str, query: str, limit: int | None = None) -> list[dict]:
    q = str(query).lower()
    rows = [r for r in _load(_resolve(sheet)) if any(q in str(v or "").lower() for v in r.values())]
    if limit and limit > 0:
        rows = rows[:limit]
    return rows


# ---------------------------------------------------------------------
# Quick self-test:  GOOGLE_API_KEY=... SPREADSHEET_IDS=... python3 data_access.py
# ---------------------------------------------------------------------
if __name__ == "__main__":
    names = list_sheet_names()
    print(f"Discovered {len(names)} sheet/tab(s): {names}")
    for s in list_sheets():
        print(" -", s)
